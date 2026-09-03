"""Gera o relatorio Excel (3 abas: Despesas Discriminadas, Resumo por Viagem,
Alertas de Consumo) a partir das viagens/despesas do periodo/empresa filtrados."""

import os
import unicodedata
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import database as db
from analise import analisar_consumo_mes, analisar_mes, analisar_viagem, categoria_chave
from utils import fmt_data, fmt_codigo, nome_motorista_padrao

FUNDO_CABECALHO = PatternFill("solid", fgColor="1F4E78")
FONTE_CABECALHO = Font(color="FFFFFF", bold=True)
FUNDO_ALERTA = PatternFill("solid", fgColor="FFC7CE")
FONTE_ALERTA = Font(color="9C0006")
FUNDO_OK = PatternFill("solid", fgColor="C6EFCE")
FONTE_OK = Font(color="006100")


def _cabecalho(ws, colunas, linha=1):
    for i, titulo in enumerate(colunas, 1):
        c = ws.cell(row=linha, column=i, value=titulo)
        c.font = FONTE_CABECALHO
        c.fill = FUNDO_CABECALHO
        c.alignment = Alignment(horizontal="center")


def _autofit(ws, larguras):
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = largura


def _aplicar_maiusculas(wb):
    """Padroniza toda célula textual do arquivo, sem alterar números ou datas."""
    for ws in wb.worksheets:
        ws.title = ws.title.upper()
        for linha in ws.iter_rows():
            for celula in linha:
                if isinstance(celula.value, str):
                    celula.value = celula.value.upper()


def _texto_padronizado(valor):
    texto = (
        unicodedata.normalize("NFKD", valor or "")
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    return "_".join(texto.strip().upper().split())


def _local_abastecimento_excel(valor):
    return {
        "ASSIS": "Assis",
        "BASE": "Assis",
        "VIAGEM": "Viagem",
        "ESTRADA": "Viagem",
    }.get(_texto_padronizado(valor), valor or "")


def _forma_pagamento_excel(valor):
    return {
        "DINHEIRO": "Dinheiro",
        "CARTAO": "Cartão",
        "PIX": "PIX",
        "FATURADO": "Faturado",
    }.get(_texto_padronizado(valor), valor or "")


def _motorista_excel(viagem):
    """Exibe nomes históricos no mesmo padrão entre MARK e ERIMAX."""
    return fmt_codigo(
        nome_motorista_padrao(viagem["empresa_nome"], viagem["motorista_nome"]),
        viagem["motorista_codigo"],
    )


def gerar_relatorio(empresa_id: int = None, caminho_saida: str = None) -> str:
    viagens = db.listar_viagens(empresa_id=empresa_id)
    despesas_todas = db.listar_despesas(empresa_id=empresa_id)
    cargas_todas = (
        db.listar_cargas_empresa(empresa_id)
        if empresa_id is not None
        else db.listar_cargas()
    )
    despesas_por_viagem = defaultdict(list)
    cargas_por_viagem = defaultdict(list)
    for despesa in despesas_todas:
        despesas_por_viagem[despesa["viagem_id"]].append(despesa)
    for carga in cargas_todas:
        cargas_por_viagem[carga["viagem_id"]].append(carga)

    wb = Workbook()

    # --- Aba 1: Despesas Discriminadas ---
    ws1 = wb.active
    ws1.title = "Despesas Discriminadas"
    colunas1 = [
        "Empresa",
        "Motorista",
        "Veiculo",
        "Viagem",
        "Categoria",
        "Data",
        "Valor (R$)",
        "Litros",
        "Local Abastecimento",
        "Forma Pagamento",
        "Descricao",
    ]
    _cabecalho(ws1, colunas1)
    linha = 2
    viagens_por_id = {v["id"]: v for v in viagens}
    for d in despesas_todas:
        v = viagens_por_id.get(d["viagem_id"])
        if v is None:
            continue
        ws1.cell(row=linha, column=1, value=v["empresa_nome"])
        ws1.cell(
            row=linha,
            column=2,
            value=_motorista_excel(v),
        )
        ws1.cell(
            row=linha,
            column=3,
            value=fmt_codigo(v["veiculo_placa"], v["veiculo_codigo"]),
        )
        ws1.cell(
            row=linha,
            column=4,
            value=f"#{v['id']} ({v['origem'] or ''} -> {v['destino'] or ''})",
        )
        ws1.cell(row=linha, column=5, value=categoria_chave(d["categoria"]))
        ws1.cell(row=linha, column=6, value=fmt_data(d["data"]))
        ws1.cell(row=linha, column=7, value=round(d["valor"], 2))
        ws1.cell(
            row=linha, column=8, value=round(d["litros"], 2) if d["litros"] else None
        )
        ws1.cell(
            row=linha,
            column=9,
            value=_local_abastecimento_excel(d["local_abastecimento"]),
        )
        ws1.cell(
            row=linha,
            column=10,
            value=_forma_pagamento_excel(d["forma_pagamento"]),
        )
        ws1.cell(row=linha, column=11, value=d["descricao"] or "")
        linha += 1
    _autofit(ws1, [14, 20, 12, 30, 14, 12, 12, 10, 16, 14, 35])

    # --- Aba 2: Cargas Discriminadas (entrega/coleta por empresa atendida) ---
    ws1b = wb.create_sheet("Cargas Discriminadas")
    colunas1b = [
        "Empresa (frota)",
        "Motorista",
        "Veiculo",
        "Viagem",
        "Empresa Atendida",
        "Tipo",
        "Data",
        "Valor (R$)",
        "Descricao",
    ]
    _cabecalho(ws1b, colunas1b)
    linha = 2
    for c in cargas_todas:
        v = viagens_por_id.get(c["viagem_id"])
        if v is None:
            continue
        ws1b.cell(row=linha, column=1, value=v["empresa_nome"])
        ws1b.cell(
            row=linha,
            column=2,
            value=_motorista_excel(v),
        )
        ws1b.cell(
            row=linha,
            column=3,
            value=fmt_codigo(v["veiculo_placa"], v["veiculo_codigo"]),
        )
        ws1b.cell(
            row=linha,
            column=4,
            value=f"#{v['id']} ({v['origem'] or ''} -> {v['destino'] or ''})",
        )
        ws1b.cell(row=linha, column=5, value=c["empresa_cliente"])
        ws1b.cell(
            row=linha, column=6, value="Entrega" if c["tipo"] == "ENTREGA" else "Coleta"
        )
        ws1b.cell(row=linha, column=7, value=fmt_data(c["data"]))
        ws1b.cell(row=linha, column=8, value=round(c["valor"], 2))
        ws1b.cell(row=linha, column=9, value=c["descricao"] or "")
        linha += 1
    _autofit(ws1b, [14, 20, 12, 30, 20, 10, 12, 12, 35])

    # --- Aba 3: Resumo por Viagem ---
    ws2 = wb.create_sheet("Resumo por Viagem")
    colunas2 = [
        "Viagem",
        "Empresa",
        "Motorista",
        "Veiculo",
        "Data Inicio",
        "Data Fim",
        "Km Rodado",
        "Litros (notas)",
        "Consumo Real (km/L)",
        "Media Painel (km/L)",
        "Diferenca % (consumo)",
        "Custo Medio Litro (R$)",
        "Preco Medio Assis (R$)",
        "Preco Medio Estrada (R$)",
        "Diferenca % (preco)",
        "Diarias",
        "Refeicoes",
        "Chapa",
        "Combustivel",
        "Arla",
        "Pecas",
        "Sinistros",
        "Pedagio",
        "Borracharia",
        "Freteiro",
        "Hospedagem",
        "Uber",
        "Outros",
        "Qtd Sinistros",
        "Total Geral (R$)",
        "Total Entrega (R$)",
        "Total Coleta (R$)",
        "Receita Total (R$)",
        "Custo sobre Receita (%)",
        "Adiantamento (R$)",
        "Devolvido (R$)",
        "Saldo a Prestar Contas (R$)",
        "Feedback Consumo",
        "Feedback Preco Combustivel",
        "Feedback Adiantamento",
    ]
    _cabecalho(ws2, colunas2)
    linha = 2
    for v in viagens:
        despesas_viagem = despesas_por_viagem[v["id"]]
        cargas_viagem = cargas_por_viagem[v["id"]]
        a = analisar_viagem(v, despesas_viagem, cargas_viagem)
        tot = a["totais_por_categoria"]
        valores = [
            v["id"],
            v["empresa_nome"],
            _motorista_excel(v),
            fmt_codigo(v["veiculo_placa"], v["veiculo_codigo"]),
            fmt_data(v["data_inicio"]),
            fmt_data(v["data_fim"]) or "(em aberto)",
            round(a["km_rodado"], 1) if a["km_rodado"] is not None else None,
            round(a["total_litros"], 2) if a["total_litros"] is not None else None,
            round(a["consumo_real"], 2) if a["consumo_real"] is not None else None,
            round(a["media_painel"], 2) if a["media_painel"] is not None else None,
            round(a["diferenca_pct"], 1) if a["diferenca_pct"] is not None else None,
            (
                round(a["custo_medio_litro"], 2)
                if a["custo_medio_litro"] is not None
                else None
            ),
            (
                round(a["preco_medio_assis"], 2)
                if a["preco_medio_assis"] is not None
                else None
            ),
            (
                round(a["preco_medio_estrada"], 2)
                if a["preco_medio_estrada"] is not None
                else None
            ),
            (
                round(a["diferenca_preco_pct"], 1)
                if a["diferenca_preco_pct"] is not None
                else None
            ),
            round(tot.get("DIARIA", 0), 2),
            round(tot.get("REFEICAO", 0), 2),
            round(tot.get("CHAPA", 0), 2),
            round(tot.get("COMBUSTIVEL", 0), 2),
            round(tot.get("ARLA", 0), 2),
            round(tot.get("PECA", 0), 2),
            round(tot.get("SINISTRO", 0), 2),
            round(tot.get("PEDAGIO", 0), 2),
            round(tot.get("BORRACHARIA", 0), 2),
            round(tot.get("FRETEIRO", 0), 2),
            round(tot.get("HOSPEDAGEM", 0), 2),
            round(tot.get("UBER", 0), 2),
            round(tot.get("OUTROS", 0), 2),
            a["qtd_sinistros"],
            round(a["total_geral"], 2),
            round(a["total_entrega"], 2),
            round(a["total_coleta"], 2),
            round(a["receita_total"], 2),
            (
                round(a["percentual_custo_receita"], 3)
                if a["percentual_custo_receita"] is not None
                else None
            ),
            round(a["adiantamento"], 2),
            round(a["devolvido"], 2),
            round(a["saldo_adiantamento"], 2),
            a["feedback"],
            a["feedback_preco"],
            a["feedback_adiantamento"],
        ]
        col_feedback_consumo = len(valores) - 2
        col_feedback_preco = len(valores) - 1
        col_feedback_adiantamento = len(valores)
        for col, val in enumerate(valores, 1):
            cel = ws2.cell(row=linha, column=col, value=val)
            if col == col_feedback_consumo:
                if a["alerta"]:
                    cel.fill = FUNDO_ALERTA
                    cel.font = FONTE_ALERTA
                elif a["consumo_real"] is not None:
                    cel.fill = FUNDO_OK
                    cel.font = FONTE_OK
            elif col == col_feedback_preco:
                if a["alerta_preco"]:
                    cel.fill = FUNDO_ALERTA
                    cel.font = FONTE_ALERTA
                elif (
                    a["preco_medio_assis"] is not None
                    and a["preco_medio_estrada"] is not None
                ):
                    cel.fill = FUNDO_OK
                    cel.font = FONTE_OK
            elif col == col_feedback_adiantamento:
                if abs(a["saldo_adiantamento"]) >= 0.01:
                    cel.fill = FUNDO_ALERTA
                    cel.font = FONTE_ALERTA
                elif a["adiantamento"] or a["devolvido"]:
                    cel.fill = FUNDO_OK
                    cel.font = FONTE_OK
        linha += 1
    _autofit(
        ws2,
        [
            8,
            12,
            18,
            10,
            12,
            12,
            10,
            12,
            16,
            16,
            14,
            16,
            16,
            16,
            14,
            10,
            10,
            10,
            12,
            10,
            10,
            10,
            10,
            12,
            10,
            12,
            10,
            10,
            10,
            14,
            14,
            14,
            14,
            16,
            14,
            14,
            16,
            45,
            45,
            45,
        ],
    )

    # --- Aba 3: Alertas de Consumo ---
    ws3 = wb.create_sheet("Alertas de Consumo")
    colunas3 = [
        "Viagem",
        "Empresa",
        "Motorista",
        "Veiculo",
        "Km Rodado",
        "Litros (notas)",
        "Consumo Real (km/L)",
        "Media Painel (km/L)",
        "Diferenca %",
        "Feedback",
    ]
    _cabecalho(ws3, colunas3)
    linha = 2
    for v in viagens:
        despesas_viagem = despesas_por_viagem[v["id"]]
        a = analisar_viagem(v, despesas_viagem)
        if not a["alerta"]:
            continue
        valores = [
            v["id"],
            v["empresa_nome"],
            _motorista_excel(v),
            fmt_codigo(v["veiculo_placa"], v["veiculo_codigo"]),
            round(a["km_rodado"], 1),
            round(a["total_litros"], 2),
            round(a["consumo_real"], 2),
            round(a["media_painel"], 2),
            round(a["diferenca_pct"], 1),
            a["feedback"],
        ]
        for col, val in enumerate(valores, 1):
            cel = ws3.cell(row=linha, column=col, value=val)
            cel.fill = FUNDO_ALERTA
            cel.font = FONTE_ALERTA
        linha += 1
    _autofit(ws3, [8, 12, 18, 10, 10, 12, 16, 16, 12, 55])
    if linha == 2:
        ws3.cell(
            row=2,
            column=1,
            value="Nenhum alerta -- todas as viagens com dados suficientes conferem com o painel.",
        )

    # --- Aba 4: Alertas de Preco Combustivel ---
    ws4 = wb.create_sheet("Alertas Preco Combustivel")
    colunas4 = [
        "Viagem",
        "Empresa",
        "Motorista",
        "Veiculo",
        "Preco Medio Assis (R$)",
        "Preco Medio Estrada (R$)",
        "Diferenca %",
        "Feedback",
    ]
    _cabecalho(ws4, colunas4)
    linha = 2
    for v in viagens:
        despesas_viagem = despesas_por_viagem[v["id"]]
        a = analisar_viagem(v, despesas_viagem)
        if not a["alerta_preco"]:
            continue
        valores = [
            v["id"],
            v["empresa_nome"],
            _motorista_excel(v),
            fmt_codigo(v["veiculo_placa"], v["veiculo_codigo"]),
            round(a["preco_medio_assis"], 2),
            round(a["preco_medio_estrada"], 2),
            round(a["diferenca_preco_pct"], 1),
            a["feedback_preco"],
        ]
        for col, val in enumerate(valores, 1):
            cel = ws4.cell(row=linha, column=col, value=val)
            cel.fill = FUNDO_ALERTA
            cel.font = FONTE_ALERTA
        linha += 1
    _autofit(ws4, [8, 12, 18, 10, 18, 18, 12, 60])
    if linha == 2:
        ws4.cell(
            row=2,
            column=1,
            value="Nenhum alerta -- o preco pago na estrada ficou dentro do praticado em Assis em todas as viagens.",
        )

    # --- Aba 5: Fechamento Mensal (entrega/coleta por empresa atendida) ---
    ws5 = wb.create_sheet("Fechamento Mensal")
    colunas5 = [
        "Mes",
        "Empresa Atendida",
        "Entrega (R$)",
        "Coleta (R$)",
        "Receita (R$)",
        "Despesa Alocada (R$)",
        "% Despesa sobre Receita",
        "Total Despesa Frete do Mes (R$)",
    ]
    _cabecalho(ws5, colunas5)
    linha = 2
    viagens_ids = {v["id"] for v in viagens}
    despesas_escopo = [d for d in despesas_todas if d["viagem_id"] in viagens_ids]
    cargas_escopo = [c for c in cargas_todas if c["viagem_id"] in viagens_ids]
    meses = sorted(
        {d["data"][:7] for d in despesas_escopo}
        | {c["data"][:7] for c in cargas_escopo}
    )
    for mes in meses:
        despesas_mes = [d for d in despesas_escopo if d["data"].startswith(mes)]
        cargas_mes = [c for c in cargas_escopo if c["data"].startswith(mes)]
        m = analisar_mes(despesas_mes, cargas_mes)
        if not m["por_empresa"]:
            ws5.cell(row=linha, column=1, value=mes)
            ws5.cell(row=linha, column=2, value="(sem cargas lancadas)")
            ws5.cell(row=linha, column=8, value=round(m["total_despesa_mes"], 2))
            linha += 1
            continue
        for emp, dd in sorted(m["por_empresa"].items(), key=lambda x: -x[1]["total"]):
            ws5.cell(row=linha, column=1, value=mes)
            ws5.cell(row=linha, column=2, value=emp)
            ws5.cell(row=linha, column=3, value=round(dd["entrega"], 2))
            ws5.cell(row=linha, column=4, value=round(dd["coleta"], 2))
            ws5.cell(row=linha, column=5, value=round(dd["total"], 2))
            ws5.cell(
                row=linha,
                column=6,
                value=(
                    round(dd["despesa_alocada"], 2)
                    if dd["despesa_alocada"] is not None
                    else None
                ),
            )
            ws5.cell(
                row=linha,
                column=7,
                value=(
                    round(dd["percentual_despesa_sobre_receita"], 3)
                    if dd["percentual_despesa_sobre_receita"] is not None
                    else None
                ),
            )
            ws5.cell(row=linha, column=8, value=round(m["total_despesa_mes"], 2))
            linha += 1
    _autofit(ws5, [10, 22, 14, 14, 14, 16, 18, 22])
    if linha == 2:
        ws5.cell(row=2, column=1, value="Nenhuma despesa ou carga lancada ainda.")

    # --- Aba 6: Consumo Mensal por Veiculo e Motorista ---
    ws6 = wb.create_sheet("Consumo Mensal")
    colunas6 = [
        "Mes",
        "Tipo",
        "Veiculo/Motorista",
        "Km Rodado",
        "Litros (notas)",
        "Consumo Medio (km/L)",
        "Qtd Viagens",
    ]
    _cabecalho(ws6, colunas6)
    linha = 2
    meses_viagens = sorted({v["data_inicio"][:7] for v in viagens})
    for mes in meses_viagens:
        viagens_mes = [v for v in viagens if v["data_inicio"].startswith(mes)]
        viagens_com_despesas = [(v, despesas_por_viagem[v["id"]]) for v in viagens_mes]
        consumo = analisar_consumo_mes(viagens_com_despesas)
        if not consumo["por_veiculo"] and not consumo["por_motorista"]:
            continue
        for placa, d in sorted(consumo["por_veiculo"].items()):
            ws6.cell(row=linha, column=1, value=mes)
            ws6.cell(row=linha, column=2, value="Veiculo")
            ws6.cell(row=linha, column=3, value=placa)
            ws6.cell(row=linha, column=4, value=round(d["km"], 1))
            ws6.cell(row=linha, column=5, value=round(d["litros"], 2))
            ws6.cell(
                row=linha,
                column=6,
                value=(
                    round(d["consumo_medio"], 2)
                    if d["consumo_medio"] is not None
                    else None
                ),
            )
            ws6.cell(row=linha, column=7, value=d["viagens"])
            linha += 1
        for nome, d in sorted(consumo["por_motorista"].items()):
            ws6.cell(row=linha, column=1, value=mes)
            ws6.cell(row=linha, column=2, value="Motorista")
            ws6.cell(row=linha, column=3, value=nome)
            ws6.cell(row=linha, column=4, value=round(d["km"], 1))
            ws6.cell(row=linha, column=5, value=round(d["litros"], 2))
            ws6.cell(
                row=linha,
                column=6,
                value=(
                    round(d["consumo_medio"], 2)
                    if d["consumo_medio"] is not None
                    else None
                ),
            )
            ws6.cell(row=linha, column=7, value=d["viagens"])
            linha += 1
    _autofit(ws6, [10, 12, 20, 14, 14, 18, 12])
    if linha == 2:
        ws6.cell(
            row=2,
            column=1,
            value="Nenhuma viagem com hodometro final e litros lancados ainda.",
        )

    if caminho_saida is None:
        base = os.path.dirname(os.path.abspath(__file__))
        nome = (
            f"relatorio_despesas_viagem_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if empresa_id is not None:
            empresa = next(
                (e for e in db.listar_empresas() if e["id"] == empresa_id), None
            )
            subpasta = empresa["nome"].lower() if empresa else "geral"
        else:
            subpasta = "todas"
        caminho_saida = os.path.join(base, "relatorios", subpasta, nome)
    os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)
    _aplicar_maiusculas(wb)
    wb.save(caminho_saida)
    return caminho_saida
